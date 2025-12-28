#!/usr/bin/env python3
"""
Generate SponsorSynq Competitor Comparison Excel Spreadsheet
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_competitor_comparison():
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # TAB 1: REVENUE STREAMS
    ws1 = wb.create_sheet("Revenue Streams")
    create_revenue_streams_tab(ws1)

    # TAB 2: QUICK COMPARISON
    ws2 = wb.create_sheet("Quick Comparison")
    create_quick_comparison_tab(ws2)

    # TAB 3: CORE DIFFERENTIATORS
    ws3 = wb.create_sheet("Core Differentiators")
    create_core_differentiators_tab(ws3)

    # TAB 4: REVENUE SUMMARY
    ws4 = wb.create_sheet("Revenue Summary")
    create_revenue_summary_tab(ws4)

    # TAB 5: ACTION ITEMS
    ws5 = wb.create_sheet("Action Items")
    create_action_items_tab(ws5)

    # Save the workbook
    wb.save("SponsorSynq_Competitor_Comparison.xlsx")
    print("✓ Excel file created: SponsorSynq_Competitor_Comparison.xlsx")


def create_revenue_streams_tab(ws):
    """Create Tab 1: Revenue Streams"""

    # Headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "Revenue Stream"
    ws['B1'] = "What We Charge"
    ws['C1'] = "What Eventbrite Charges"
    ws['D1'] = "What Others Charge"
    ws['E1'] = "Where We Win"
    ws['F1'] = "Where We Need to Improve"

    for col in range(1, 7):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Revenue Stream Data
    streams = [
        {
            "name": "1. Processing Fee",
            "we_charge": "~3% + $0.30 per ticket\nPassed to attendee\nStandard, expected",
            "eventbrite": "3.7% + $1.79 service fee\nPLUS 2.9% processing\nTotal: ~6.6% + $1.79\nOn $20 ticket: $3.11",
            "others": "Ticketmaster: 15-25%\nTicket Tailor: $0.26 flat\nHumanitix: 2% + $0.30",
            "win": "• Significantly cheaper than Eventbrite & Ticketmaster\n• Transparent pricing\n• No hidden fees",
            "improve": "• Don't compete on fees alone\n• Our story: 'Pay $50 in fees, get $500 in sponsors'\n• Emphasize net gain, not fee savings"
        },
        {
            "name": "2. Platform Fee",
            "we_charge": "5% of ticket revenue\nOnly on 2nd+ events\n1st event FREE\nWaived with $19/mo subscription",
            "eventbrite": "~6.6% + $1.79 per ticket\nChanged pricing 11 times\nRemoved free tier in 2023\nConfusing structure",
            "others": "Meetup: $16-22/mo required\nFacebook Events: Free (no ticketing)\nEventcube: Similar to us",
            "win": "• Simple & transparent\n• Cheaper than Eventbrite\n• First event free builds trust\n• Subscription waives entirely",
            "improve": "• Publicly commit to pricing stability\n• Never surprise users like Eventbrite did\n• Emphasize 'pay only when you win'"
        },
        {
            "name": "3. Monthly Subscription",
            "we_charge": "$19/month for Pro\nWaives 5% fee entirely\nIncludes Premium Analytics ($7 value)\nBreakeven: $380/mo in sales",
            "eventbrite": "Pro plans: $15-100/month\nFor email marketing limits\nDoes NOT waive fees\nYou pay subscription AND fees",
            "others": "Meetup: $16-22/mo required\nNo free tier\nPay even if event flops",
            "win": "• Subscription SAVES money\n• Eventbrite double-dips\n• Meetup charges even with no revenue\n• We reward active hosts",
            "improve": "• Offer annual discount ($190/yr vs $228)\n• Show savings in dashboard monthly\n• Turn subscription into psychological win"
        },
        {
            "name": "4. Sponsor Commission",
            "we_charge": "12% of sponsor allocations\nTaken before host payout\n$100 = $12 to us, $88 to host\nAutomatic, no invoice",
            "eventbrite": "NOTHING\nNo sponsor marketplace exists\nHosts find sponsors alone\nCold emails, agencies, luck",
            "others": "Agencies: 15-25% + retainers\nInfluencer platforms: 10-20%\nTicketmaster: No marketplace",
            "win": "• ONLY platform with sponsor marketplace\n• Eventbrite can't compete here\n• 12% vs 15-25% agencies\n• Accessible to any event size",
            "improve": "• Matching algorithm must be excellent\n• Verification system must be robust\n• Show ROI metrics to sponsors\n• Prove 12% is worth it"
        },
        {
            "name": "5. Featured Placement",
            "we_charge": "$29 flat fee\n14 days priority placement\nSponsors appear first\n'Featured' badge",
            "eventbrite": "Eventbrite Ads: CPC/CPM model\n$50-500+ campaigns\nFor events→attendees\nNot sponsors→hosts",
            "others": "LinkedIn: $30-100 boost\nIndeed: $5-25/day\nInstagram: $1+/day",
            "win": "• Simple flat pricing\n• No bidding or budgets\n• Specifically for sponsor-host connections\n• Different use case than competitors",
            "improve": "• Test price tiers ($15/7d, $29/14d, $49/30d)\n• Track & display performance\n• Prove Featured = more acceptances"
        },
        {
            "name": "6. Instant Payout",
            "we_charge": "1.5% of payout amount\nSame-day deposit\nStandard (3-5 days) is free\nCompletely optional",
            "eventbrite": "Standard: 5 business days\nNO instant option\nNeed money faster? Too bad.",
            "others": "Stripe direct: 1%\nDoorDash: $1.99 flat\nUber: $0.85\nPayPal: 1.75%",
            "win": "• We offer it - Eventbrite doesn't\n• Competitive with PayPal\n• Meaningful differentiation",
            "improve": "• Consider flat fee for large payouts\n• Implement cap: 1.5% max $15\n• Communicate value: 'Pay DJ tonight'"
        },
        {
            "name": "7. Premium Analytics",
            "we_charge": "$7/month standalone\nFree with Pro subscription\nTraffic sources, timing, feedback\nExportable reports",
            "eventbrite": "Basic: Free\nAdvanced: Bundled in $15-100/mo Pro\nCan't buy separately",
            "others": "Mixpanel: $20+/mo (general tool)\nSplash: $100s/mo (enterprise)\nGoogle Analytics: Free (not event-specific)",
            "win": "• Event-specific at low price\n• $7/mo accessible to all\n• Free with subscription adds value",
            "improve": "• Make insights actionable\n• Connect to recommended actions\n• Add sponsor-specific analytics\n• Create feedback loop"
        },
        {
            "name": "8. Event Boost",
            "we_charge": "$15 for 7 days\nPriority in sponsor discovery\nCompletely optional",
            "eventbrite": "Eventbrite Ads: $50-500+\nVariable CPC/CPM\nFor events→attendees only",
            "others": "Instagram boost: $7-70/week\nFacebook boost: ~$14/week\nBoth for attendees, not sponsors",
            "win": "• Designed to attract SPONSORS\n• Unique product\n• $15 boost → sponsor offers\n• Direct ROI vs random reach",
            "improve": "• Track & display performance\n• Show if boost → more offers\n• Offer tiers: $10/3d, $15/7d, $25/14d"
        },
        {
            "name": "9. Enterprise Licensing",
            "we_charge": "$10K-25K/year\nCustom branding\nAdmin dashboards\nCompliance features",
            "eventbrite": "Custom pricing\n$10Ks-$100Ks/year\nFocus: Corporations, festivals\nNot designed for universities",
            "others": "Presence: $10K-50K/yr (no ticketing)\nCampusLabs: $15K-75K/yr (no sponsors)\nMeetup Pro: $2.5K+/yr",
            "win": "• All-in-one: ticketing + sponsors + management\n• Others need multiple tools\n• Designed for student events",
            "improve": "• Build admin/compliance features\n• Approval workflows, audit trails\n• SIS integration\n• Develop case studies"
        },
        {
            "name": "10. Ambassador Program",
            "we_charge": "Tiered rewards:\n• Starter: $15/referral\n• Rising: $20 + free Analytics\n• Elite: $25 + 1% rev share\n• Founding: $30 + 2% rev share",
            "eventbrite": "Give $10, Get $10\nCaps at $50 total\nNo tiers\nNo revenue share\nOne-time only",
            "others": "Uber: $5-20 one-time\nDropbox: Storage bonus\nAirbnb: Travel credit\nAll one-time, no ongoing",
            "win": "• Tiered = gamification\n• Revenue share = ongoing stake\n• No cap on earnings\n• Ambassadors help referrals succeed",
            "improve": "• Surface program prominently\n• Show earnings on dashboard\n• Create leaderboards\n• Make Founding feel like achievement"
        }
    ]

    row = 2
    for stream in streams:
        ws[f'A{row}'] = stream['name']
        ws[f'B{row}'] = stream['we_charge']
        ws[f'C{row}'] = stream['eventbrite']
        ws[f'D{row}'] = stream['others']
        ws[f'E{row}'] = stream['win']
        ws[f'F{row}'] = stream['improve']

        # Style the row
        name_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        ws[f'A{row}'].fill = name_fill
        ws[f'A{row}'].font = Font(bold=True, size=11)

        for col in range(1, 7):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35

    # Set row heights
    for i in range(2, row):
        ws.row_dimensions[i].height = 100


def create_quick_comparison_tab(ws):
    """Create Tab 2: Quick Comparison"""

    # Headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "Feature"
    ws['B1'] = "SponsorSynq"
    ws['C1'] = "Eventbrite"
    ws['D1'] = "Winner"

    for col in range(1, 5):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Comparison data
    comparisons = [
        ["Processing Fee", "~3% + $0.30", "~6.6% + $1.79", "SponsorSynq"],
        ["Platform Fee", "5% (waivable)", "~6.6% + $1.79 (not waivable)", "SponsorSynq"],
        ["First Event", "FREE", "Fees apply", "SponsorSynq"],
        ["Subscription Waives Fees?", "YES", "NO (double-dip)", "SponsorSynq"],
        ["Sponsor Marketplace", "YES (built-in)", "NO", "SponsorSynq"],
        ["Sponsor Matching AI", "YES", "NO", "SponsorSynq"],
        ["Instant Payout", "YES (1.5%)", "NO", "SponsorSynq"],
        ["Pricing Changes Since 2007", "0 (we're new)", "11 times", "SponsorSynq"],
        ["Referral Program Cap", "UNLIMITED + rev share", "$50 max", "SponsorSynq"],
        ["Free Event Posting", "YES", "YES (restored after backlash)", "Tie"],
        ["Brand Recognition", "New/Unknown", "Industry Leader", "Eventbrite"],
        ["Enterprise Clients", "Building", "Established", "Eventbrite"],
        ["Global Infrastructure", "Building", "180+ countries", "Eventbrite"]
    ]

    row = 2
    for comp in comparisons:
        ws[f'A{row}'] = comp[0]
        ws[f'B{row}'] = comp[1]
        ws[f'C{row}'] = comp[2]
        ws[f'D{row}'] = comp[3]

        # Color code winners
        winner_fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        winner_fill_yellow = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        winner_fill_red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        if comp[3] == "SponsorSynq":
            ws[f'D{row}'].fill = winner_fill_green
            ws[f'D{row}'].font = Font(bold=True, color="006100")
        elif comp[3] == "Tie":
            ws[f'D{row}'].fill = winner_fill_yellow
            ws[f'D{row}'].font = Font(bold=True, color="9C6500")
        elif comp[3] == "Eventbrite":
            ws[f'D{row}'].fill = winner_fill_red
            ws[f'D{row}'].font = Font(bold=True, color="9C0006")

        for col in range(1, 5):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 20


def create_core_differentiators_tab(ws):
    """Create Tab 3: Core Differentiators"""

    # Headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "Differentiator"
    ws['B1'] = "What It Means"
    ws['C1'] = "Why It Matters"

    for col in range(1, 4):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Differentiator data
    differentiators = [
        {
            "name": "1. SPONSOR MARKETPLACE",
            "means": "We have one. They don't.",
            "matters": "• Hosts make money they'd never find otherwise\n• This is our #1 differentiator\n• Eventbrite literally cannot compete"
        },
        {
            "name": "2. REVENUE PLATFORM vs TICKETING PLATFORM",
            "means": "We help hosts MAKE money\nThey just help hosts SELL tickets",
            "matters": "• Different value proposition entirely\n• We're not a cheaper Eventbrite\n• We're a different category"
        },
        {
            "name": "3. SUBSCRIPTION ACTUALLY SAVES MONEY",
            "means": "Our $19/mo waives fees\nTheir $15-100/mo doesn't",
            "matters": "• Eventbrite double-dips\n• We reward active hosts\n• Clear economic value"
        },
        {
            "name": "4. FIRST EVENT FREE",
            "means": "Zero risk to try\nExperience value before paying",
            "matters": "• Builds trust\n• Removes 'what if it doesn't work' anxiety\n• Converts skeptics"
        },
        {
            "name": "5. PRICING STABILITY",
            "means": "We won't surprise users\nThey've changed 11 times",
            "matters": "• Eventbrite destroyed trust\n• We can win on reliability\n• Public commitment matters"
        },
        {
            "name": "6. AMBASSADOR PROGRAM",
            "means": "Tiered rewards + revenue share\nTheirs caps at $50",
            "matters": "• Creates true evangelists\n• Ongoing stake in platform success\n• Organic growth engine"
        }
    ]

    row = 2
    for diff in differentiators:
        ws[f'A{row}'] = diff['name']
        ws[f'B{row}'] = diff['means']
        ws[f'C{row}'] = diff['matters']

        # Style
        name_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{row}'].fill = name_fill
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")

        for col in range(1, 4):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45

    # Set row heights
    for i in range(2, row):
        ws.row_dimensions[i].height = 75


def create_revenue_summary_tab(ws):
    """Create Tab 4: Revenue Summary"""

    # Headers
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)

    ws['A1'] = "Stream"
    ws['B1'] = "Who Pays"
    ws['C1'] = "Amount"
    ws['D1'] = "When It Triggers"
    ws['E1'] = "Competitors Have This?"

    for col in range(1, 6):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Revenue summary data
    revenues = [
        ["Processing Fee", "Attendee", "~3% + $0.30", "Every ticket", "YES - but higher"],
        ["Platform Fee", "Host", "5%", "2nd+ event, non-subscriber", "YES - but not waivable"],
        ["Subscription", "Host", "$19/month", "Optional, waives 5% fee", "YES - but doesn't waive fees"],
        ["Sponsor Commission", "Sponsor", "12%", "Every sponsorship deal", "NO - we're the only one"],
        ["Featured Placement", "Sponsor", "$29", "Optional add-on", "NO - unique product"],
        ["Instant Payout", "Host", "1.5%", "Optional, on-demand", "NO - Eventbrite doesn't offer"],
        ["Premium Analytics", "Host", "$7/month", "Optional (free w/ Pro)", "PARTIAL - bundled only"],
        ["Event Boost", "Host", "$15", "Optional, per event", "NO - unique to sponsors"],
        ["Enterprise License", "University", "$10K-25K/year", "Annual contract", "YES - but no sponsor integration"],
        ["Ambassador Program", "N/A (we pay)", "Credits + rev share", "Successful referrals", "WEAK - Eventbrite caps at $50"]
    ]

    row = 2
    for rev in revenues:
        ws[f'A{row}'] = rev[0]
        ws[f'B{row}'] = rev[1]
        ws[f'C{row}'] = rev[2]
        ws[f'D{row}'] = rev[3]
        ws[f'E{row}'] = rev[4]

        # Highlight unique offerings
        unique_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        if "NO - " in rev[4] or "unique" in rev[4].lower():
            ws[f'E{row}'].fill = unique_fill
            ws[f'E{row}'].font = Font(bold=True, color="006100")

        for col in range(1, 6):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35


def create_action_items_tab(ws):
    """Create Tab 5: Action Items"""

    # Title
    title_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    title_font = Font(color="FFFFFF", bold=True, size=14)

    ws.merge_cells('A1:D1')
    ws['A1'] = "THE CORE MESSAGE"
    ws['A1'].fill = title_fill
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A2:D2')
    ws['A2'] = "We are NOT a cheaper Eventbrite."
    ws['A2'].font = Font(bold=True, size=12, color="C00000")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A3:D3')
    ws['A3'] = "We are a REVENUE GENERATION platform."
    ws['A3'].font = Font(bold=True, size=12, color="C00000")
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A4:D4')
    ws['A4'] = "Eventbrite is just a TICKETING platform."
    ws['A4'].font = Font(bold=True, size=11)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A5:D5')
    ws['A5'] = "Hosts who use us make more money because sponsors find them automatically."
    ws['A5'].font = Font(bold=True, size=11)
    ws['A5'].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells('A6:D6')
    ws['A6'] = "That's the story. That's how we win."
    ws['A6'].font = Font(bold=True, size=12, color="C00000")
    ws['A6'].alignment = Alignment(horizontal="center", vertical="center")

    # Spacing
    ws.row_dimensions[7].height = 5

    # HIGH PRIORITY Section
    high_header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    high_header_font = Font(color="FFFFFF", bold=True, size=12)

    ws.merge_cells('A8:D8')
    ws['A8'] = "HIGH PRIORITY"
    ws['A8'].fill = high_header_fill
    ws['A8'].font = high_header_font
    ws['A8'].alignment = Alignment(horizontal="center", vertical="center")

    ws['A9'] = "Area"
    ws['B9'] = "What To Build/Do"
    ws['C9'] = "Why It Matters"
    ws['D9'] = "Priority"

    for col in range(1, 5):
        cell = ws.cell(9, col)
        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    high_priority = [
        ["Event Collaboration/Co-hosting", "Multi-host event collaboration system:\n• Primary & secondary event hosts\n• Real-time ticket revenue tracking for all hosts\n• Customizable revenue split (50/50, 60/40, etc.)\n• Primary host controls fund distribution\n• Venue payment tracking/metrics\n• Multi-city tour dashboard\n• Track all events under one tour umbrella", "Game-changer for large-scale events and tours. Creates network effects - hosts bring other hosts. Differentiates us from all competitors who only support single-host events.", "HIGH"],
        ["Promoter Referral System", "Role-based promoter program:\n• Event hosts assign 'promoter' roles\n• Unique tracking links/QR codes per promoter\n• Track ticket sales by promoter link\n• Automatic commission payouts\n• Works like affiliate program but for event promotion\n• Promoters share on Instagram/TikTok stories", "Viral growth engine. Turns every event into distributed sales team. Promoters are incentivized to drive ticket sales. Event hosts get free marketing.", "HIGH"],
        ["Landing Page Overhaul", "PRICING: Only 2 tiers:\n• Free: Unlimited events, basic features\n• $100/mo: Money-making features (was $19)\n\nMESSAGING:\n• Focus on ROI, not features\n• 'Save X, Make Y' psychology\n• Fewer words, clearer value\n• Show savings/earnings potential per feature\n\nCOMPETITOR COMPARISON:\n• REMOVE from landing page entirely\n• Move to post-signup dashboard\n• Only show after they're already users\n• Prevent competitors from copying our strategy", "Simplicity converts better. People don't want subscriptions - they want to 'pay when they get paid' or see clear ROI. Hiding competitor intel protects our competitive advantage.", "HIGH"],
        ["Venue Partnership System", "Venue integration features:\n• Venues can create accounts\n• Event hosts connect their venue\n• Transparent revenue sharing dashboard\n• Track bar/food/door splits in real-time\n• Simple fund distribution\n\nVENUE LOCK-IN STRATEGY:\n• Partner with venues to require SponsorSynq\n• 'Want to use our venue? Use SponsorSynq'\n• Creates vendor lock-in\n• Eliminates competition with Eventbrite/Posh\n• Market domination through venue partnerships", "This is the ULTIMATE competitive moat. If venues require SponsorSynq, event hosts have no choice. We stop competing on features and own the distribution channel. Venue adoption = market domination.", "HIGH"],
        ["Revenue Stream Documentation", "Comprehensive revenue analysis:\n• Document ALL current revenue streams\n• Platform fee, processing, sponsorship commission, etc.\n• Strategy for free events (no ticket sales)\n• How to monetize users who don't charge for tickets?\n• Options: Sponsorship-only commission, require subscription, tiered free limits\n• Question: Is free platform usage okay if it brings brand awareness?\n• Alternative revenue from free event hosts?", "Need clarity on business model edge cases. Free events still have value (brand awareness, sponsor discovery) but need strategy to ensure sustainable revenue. Must balance growth with monetization.", "HIGH"]
    ]

    row = 10
    for item in high_priority:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'C{row}'] = item[2]
        ws[f'D{row}'] = item[3]

        for col in range(1, 5):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set row heights for high priority items (more content now)
    for i in range(10, row):
        ws.row_dimensions[i].height = 120

    # Spacing
    ws.row_dimensions[row].height = 5
    row += 1

    # MEDIUM PRIORITY Section
    med_header_fill = PatternFill(start_color="FF9900", end_color="FF9900", fill_type="solid")
    med_header_font = Font(color="FFFFFF", bold=True, size=12)

    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "MEDIUM PRIORITY"
    ws[f'A{row}'].fill = med_header_fill
    ws[f'A{row}'].font = med_header_font
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f'A{row}'] = "Area"
    ws[f'B{row}'] = "What To Build/Do"
    ws[f'C{row}'] = "Why It Matters"
    ws[f'D{row}'] = "Priority"

    for col in range(1, 5):
        cell = ws.cell(row, col)
        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    medium_priority = [
        ["Performance Tracking", "Track Boost & Featured Placement results", "Need data to prove these upgrades work", "MEDIUM"],
        ["Instant Payout Cap", "Implement 1.5% with $15 max", "Makes instant payout attractive for large payouts", "MEDIUM"],
        ["Annual Subscription", "Offer $190/year option", "Locks in committed hosts, improves predictability", "MEDIUM"],
        ["Tiered Boost Options", "$10/3d, $15/7d, $25/14d", "Flexibility for different event timelines", "MEDIUM"]
    ]

    for item in medium_priority:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'C{row}'] = item[2]
        ws[f'D{row}'] = item[3]

        for col in range(1, 5):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Spacing
    ws.row_dimensions[row].height = 5
    row += 1

    # LOW PRIORITY Section
    low_header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    low_header_font = Font(color="FFFFFF", bold=True, size=12)

    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = "LOW PRIORITY"
    ws[f'A{row}'].fill = low_header_fill
    ws[f'A{row}'].font = low_header_font
    ws[f'A{row}'].alignment = Alignment(horizontal="center", vertical="center")

    row += 1
    ws[f'A{row}'] = "Area"
    ws[f'B{row}'] = "What To Build/Do"
    ws[f'C{row}'] = "Why It Matters"
    ws[f'D{row}'] = "Priority"

    for col in range(1, 5):
        cell = ws.cell(row, col)
        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    row += 1
    low_priority = [
        ["Enterprise Compliance", "Approval workflows, audit trails, SIS integration", "Required for university sales", "LOW"],
        ["Leaderboards", "Show top ambassadors", "Creates competition, shows what's possible", "LOW"]
    ]

    for item in low_priority:
        ws[f'A{row}'] = item[0]
        ws[f'B{row}'] = item[1]
        ws[f'C{row}'] = item[2]
        ws[f'D{row}'] = item[3]

        for col in range(1, 5):
            cell = ws.cell(row, col)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 12


if __name__ == "__main__":
    create_competitor_comparison()
